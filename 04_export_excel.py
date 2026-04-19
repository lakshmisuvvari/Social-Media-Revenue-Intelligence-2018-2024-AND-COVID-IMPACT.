"""
04_export_excel.py
==================
Exports a styled multi-sheet Excel workbook from the clean CSV
and the 5 SQL query result CSVs.

Usage:
    python 04_export_excel.py

Output:
    ../output/social_media_report.xlsx
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import os

# ── Paths ──────────────────────────────────────────────────────────
OUTPUT_DIR  = "../output"
CLEAN_CSV   = f"{OUTPUT_DIR}/clean_social_media.csv"
OUT_XLSX    = f"{OUTPUT_DIR}/social_media_report.xlsx"

# ── Colour Palette ─────────────────────────────────────────────────
DARK_BLUE   = "1F3864"   # header background
LIGHT_BLUE  = "BDD7EE"   # sub-header / alternating row
WHITE       = "FFFFFF"
ACCENT_RED  = "C00000"   # decline values
ACCENT_GRN  = "375623"   # growth values
ALT_ROW     = "EBF3FB"   # alternating row tint

# ── Style Helpers ──────────────────────────────────────────────────
def header_style(cell, bg=DARK_BLUE, fg=WHITE, size=11, bold=True):
    cell.font      = Font(name="Arial", bold=bold, color=fg, size=size)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def write_sheet(wb, title, df, freeze="A2", tab_color=DARK_BLUE,
                pct_cols=None, money_cols=None):
    """Write a DataFrame to a styled worksheet."""
    ws = wb.create_sheet(title=title)
    ws.tab_color = tab_color

    # ── Title row ──────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=len(df.columns))
    title_cell = ws.cell(row=1, column=1, value=title)
    header_style(title_cell, bg=DARK_BLUE, size=13)
    ws.row_dimensions[1].height = 28

    # ── Column headers ─────────────────────────────────────────────
    for c_idx, col in enumerate(df.columns, 1):
        cell = ws.cell(row=2, column=c_idx,
                       value=col.replace("_", " ").title())
        header_style(cell, bg="2E75B6", size=10)
    ws.row_dimensions[2].height = 22

    # ── Data rows ──────────────────────────────────────────────────
    for r_idx, row in enumerate(df.itertuples(index=False), start=3):
        bg = ALT_ROW if r_idx % 2 == 0 else WHITE
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = PatternFill("solid", start_color=bg)
            cell.border    = thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")

            col_name = df.columns[c_idx - 1]

            # Money columns → $#,##0.0
            if money_cols and col_name in money_cols and isinstance(value, (int, float)):
                cell.number_format = '#,##0.0'

            # Percent columns → colour + format
            if pct_cols and col_name in pct_cols and isinstance(value, (int, float)):
                cell.number_format = '0.0"%"'
                if value < 0:
                    cell.font = Font(name="Arial", size=10,
                                     bold=True, color=ACCENT_RED)
                elif value > 50:
                    cell.font = Font(name="Arial", size=10,
                                     bold=True, color=ACCENT_GRN)

    # ── Column widths ──────────────────────────────────────────────
    for c_idx, col in enumerate(df.columns, 1):
        max_len = max(
            len(str(col).replace("_", " ").title()),
            df[col].astype(str).str.len().max() if len(df) > 0 else 0
        )
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 4, 30)

    # ── Freeze panes ───────────────────────────────────────────────
    ws.freeze_panes = freeze
    return ws


# ── Load data ──────────────────────────────────────────────────────
print("=" * 55)
print("STEP 1 — Loading data files")
print("=" * 55)

clean   = pd.read_csv(CLEAN_CSV)
q1      = pd.read_csv(f"{OUTPUT_DIR}/q1_total_revenue_by_platform.csv")
q2      = pd.read_csv(f"{OUTPUT_DIR}/q2_revenue_by_era.csv")
q3      = pd.read_csv(f"{OUTPUT_DIR}/q3_top_growth_spurts.csv")
q4      = pd.read_csv(f"{OUTPUT_DIR}/q4_decline_years.csv")
q5      = pd.read_csv(f"{OUTPUT_DIR}/q5_industry_totals_by_year.csv")
print(f"  Clean data  : {len(clean)} rows")
print(f"  Query files : 5 loaded")

# ── Build workbook ─────────────────────────────────────────────────
print("\nSTEP 2 — Building Excel workbook")

wb = Workbook()
wb.remove(wb.active)   # remove default blank sheet

# Sheet 1 — Full clean dataset
write_sheet(wb, "📊 Full Dataset", clean,
            freeze="A3",
            tab_color="1F3864",
            pct_cols=["yoy_revenue_growth_pct", "ad_revenue_share_pct"],
            money_cols=["revenue_usd_billion", "ad_revenue_usd_billion",
                        "other_revenue_usd_billion", "revenue_usd_million"])
print("  ✅ Sheet 1 — Full Dataset")

# Sheet 2 — Q1 Total revenue by platform
write_sheet(wb, "Q1 · Revenue by Platform", q1,
            tab_color="2E75B6",
            pct_cols=["avg_yoy_growth_pct"],
            money_cols=["total_revenue_b", "min_revenue_b", "max_revenue_b"])
print("  ✅ Sheet 2 — Q1 Revenue by Platform")

# Sheet 3 — Q2 Era comparison
write_sheet(wb, "Q2 · COVID Era Impact", q2,
            tab_color="70AD47",
            money_cols=["avg_pre_covid", "avg_covid",
                        "avg_recovery", "avg_post_covid"])
print("  ✅ Sheet 3 — Q2 COVID Era Impact")

# Sheet 4 — Q3 Top growth spurts
write_sheet(wb, "Q3 · Growth Leaderboard", q3,
            tab_color="FFC000",
            pct_cols=["yoy_pct"],
            money_cols=["revenue_b"])
print("  ✅ Sheet 4 — Q3 Growth Leaderboard")

# Sheet 5 — Q4 Decline years
write_sheet(wb, "Q4 · Decline Years", q4,
            tab_color="C00000",
            pct_cols=["yoy_pct"],
            money_cols=["revenue_b"])
print("  ✅ Sheet 5 — Q4 Decline Years")

# Sheet 6 — Q5 Industry totals
write_sheet(wb, "Q5 · Industry Totals", q5,
            tab_color="7030A0",
            pct_cols=["avg_growth_pct"],
            money_cols=["total_industry_b", "total_ad_revenue_b"])
print("  ✅ Sheet 6 — Q5 Industry Totals")

# ── Summary / Cover sheet (insert first) ──────────────────────────
cover = wb.create_sheet(title="📋 Summary", index=0)
cover.tab_color = "1F3864"
cover.column_dimensions["A"].width = 35
cover.column_dimensions["B"].width = 30

summary_data = [
    ("Social Media Revenue & COVID Impact", ""),
    ("", ""),
    ("Project",        "Social Media Revenue Analysis"),
    ("Data Range",     "2018 – 2024"),
    ("Platforms",      "7  (Facebook, YouTube, Instagram, TikTok, Twitter, Snapchat, LinkedIn)"),
    ("Total Rows",     f"{len(clean)}"),
    ("Sheets",         "6  (Full data + 5 analyst queries)"),
    ("", ""),
    ("KEY FINDINGS", ""),
    ("Highest Total Revenue",    "Facebook  $738B"),
    ("Fastest Grower",           "TikTok  +400% in 2019"),
    ("Worst Decline",            "Twitter  -31.8% in 2023"),
    ("Industry Size 2018",       "$89.6B across 7 platforms"),
    ("Industry Size 2024",       "$338.4B across 7 platforms"),
    ("Ad Revenue Share 2024",    "~88% of total industry revenue"),
]

for r, (label, value) in enumerate(summary_data, 1):
    lc = cover.cell(row=r, column=1, value=label)
    vc = cover.cell(row=r, column=2, value=value)

    if r == 1:
        lc.font = Font(name="Arial", bold=True, size=16, color=WHITE)
        lc.fill = PatternFill("solid", start_color=DARK_BLUE)
        lc.alignment = Alignment(horizontal="left", vertical="center")
        cover.row_dimensions[r].height = 32
        cover.merge_cells(f"A1:B1")
    elif label == "KEY FINDINGS":
        for cell in [lc, vc]:
            cell.font = Font(name="Arial", bold=True, size=11, color=WHITE)
            cell.fill = PatternFill("solid", start_color="2E75B6")
            cell.alignment = Alignment(horizontal="left", vertical="center")
        cover.row_dimensions[r].height = 20
    elif label:
        lc.font = Font(name="Arial", bold=True, size=10)
        vc.font = Font(name="Arial", size=10)
        lc.fill = PatternFill("solid", start_color=ALT_ROW)
        vc.fill = PatternFill("solid", start_color=ALT_ROW)
        for cell in [lc, vc]:
            cell.border    = thin_border()
            cell.alignment = Alignment(horizontal="left",
                                       vertical="center", wrap_text=True)
        cover.row_dimensions[r].height = 18

print("  ✅ Cover — Summary sheet")

# ── Save ───────────────────────────────────────────────────────────
os.makedirs(OUTPUT_DIR, exist_ok=True)
wb.save(OUT_XLSX)

print(f"\n{'=' * 55}")
print("✅  COMPLETE")
print(f"{'=' * 55}")
print(f"  Saved → {OUT_XLSX}")
print(f"  Sheets : {len(wb.sheetnames)}")
for s in wb.sheetnames:
    print(f"    • {s}")
print("\nNext → Build Power BI dashboard")